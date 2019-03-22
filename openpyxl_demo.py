from openpyxl import load_workbook 

file = r"C:\Users\Supegg\Desktop\test呵呵.xlsx"
wb = load_workbook(file)
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet["C"])    # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
print(sheet["4"])    # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
print(type(sheet["C4"].value))    # c4     <-第C4格的值
print(sheet.max_row)    # 10     <-最大行数
print(sheet.max_column)    # 5     <-最大列数
for i in sheet["C"]:
    i.value += "呵呵"
    print(i.value, end=" ")
    
wb.save(file)